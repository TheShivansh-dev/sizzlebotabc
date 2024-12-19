# Token and Bot Username
#TOKEN: Final = '7673671830:AAFaDzia9GXrXAz86UEFwzkXGB7OUEFb3xM'
#BOT_USERNAME: Final = '@slizyy_bot'
#TOKEN: Final = '7007935023:AAENkGaklw6LMJA_sfhVZhnoAgIjW4lDTBc'
#BOT_USERNAME: Final = '@Grovieee_bot'
#ALLOWED_GROUP_IDS = [-1001817635995, -1002114430690]

import os
import random
import re
import pandas as pd
import openpyxl
from typing import Final
from telegram import Update, PollAnswer, Poll, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, PollAnswerHandler, CallbackQueryHandler, ContextTypes
from collections import defaultdict
import asyncio
from openpyxl import load_workbook, Workbook
import time
from telegram.error import Forbidden,BadRequest, TimedOut
import telegram

# Bot configuration
TOKEN: Final = '7673671830:AAFaDzia9GXrXAz86UEFwzkXGB7OUEFb3xM'
BOT_USERNAME: Final = '@slizyy_bot'
ALLOWED_GROUP_IDS = [-1001817635995, -1002114430690,-1002359766306]
EXCEL_FILE = 'SYNO5.xlsx'
SCORE_FILE="user_scores.xlsx"



# Global state variables

quiz_state = {}
correct_users = {}  # Tracks correct answers per user
selected_poll_count = 0 
selected_quizscore_count=0
active_poll=1 # Number of polls user requested
answers_received = defaultdict(int)  # Tracks how many answers have been received for each user
is_quiz_active = False  # New variable to track if a quiz is active
chat_id = None  # Current chat ID for the quiz
selected_time_limit = 10  # Default time limit
unanswered_poll = 0
cancel_active = False
display_chat=0
Quiz_grammar_type =''
quiz_kick= False

# Load quiz data from Excel
used_srnos = set()
def reset_used_srnos():
    global used_srnos
    used_srnos.clear()
def load_quiz_data(file_path, selected_poll_count):
    global used_srnos
    try:
        df = pd.read_excel(file_path)
        
        # Trim extra spaces in each column where cells are strings
        for col in df.select_dtypes(include="object").columns:
            df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
        
        # Filter out rows that have already been selected based on `srno`
        unique_rows = df[~df['srno'].isin(used_srnos)]
        
        # If there are fewer unique rows than requested polls, adjust to available rows
        if len(unique_rows) < selected_poll_count:
            print("Not enough unique rows available.")
            selected_poll_count = len(unique_rows)
        
        # Select a random sample of unique rows
        selected_rows = unique_rows.sample(n=selected_poll_count)
        
        # Update used_srnos with newly selected rows
        used_srnos.update(selected_rows['srno'].tolist())
        
        # Process selected rows into polls
        polls = []
        for _, row in selected_rows.iterrows():
            options = [row["option1"], row["option2"], row["option3"], row["option4"]]
            random.shuffle(options) 
            poll = {
                "question": row["question"],
                "options": options,
                "correct_answer": row["answer"],
                "meaning": row.get("meaning", "No meaning provided")  # Ensure meaning is loaded
            }
            polls.append(poll)
            
        return polls
    except Exception as e:
        print(e)
def escape_markdown(text: str) -> str:
    return re.sub(r'([_\*\[\]\(\)~`>#+\-=|{}.!])', r'\\\1', text)
async def delete_user_scores(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Deletes all rows in the userscore table except the headers.
    """
    try:
        # Load the workbook and sheet
        workbook = load_workbook(SCORE_FILE)
        sheet = workbook.active

        # Check if there are rows to delete (beyond the header)
        if sheet.max_row > 1:
            # Delete all rows except the header
            sheet.delete_rows(2, sheet.max_row - 1)

            # Save the updated workbook
            workbook.save(SCORE_FILE)

            await update.message.reply_text("All user scores have been deleted successfully.")
        else:
            await update.message.reply_text("The user score table is already empty.")

    except FileNotFoundError:
        await update.message.reply_text("The score file does not exist. No action was taken.")
    except Exception as e:
        await update.message.reply_text(f"An error occurred while deleting scores: {e}")
def load_scores():
    if not os.path.exists(SCORE_FILE):
        return []

    workbook = openpyxl.load_workbook(SCORE_FILE)
    sheet = workbook.active

    scores = []
    for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip the header

        user_id = sheet.cell(row=row, column=2).value
        username = sheet.cell(row=row, column=3).value
        score = sheet.cell(row=row, column=4).value
        round = sheet.cell(row=row, column=5).value

        if user_id and username and score and round is not None:
            scores.append((user_id, username, score,round))

    workbook.close()
    return scores

async def select_top_10_users(update: Update, context: ContextTypes.DEFAULT_TYPE):
    scores = load_scores()
    if not scores:
        try:
            await update.message.reply_text("No scores found")
        except telegram.error.BadRequest:
            await update.message.chat.send_message("No scores found")
        return
    # Sort by score in descending order
    scores.sort(key=lambda x: x[2], reverse=True)
    # Get the top 10 users
    top_10 = scores[:10]

    # Build the message to display top users
    message = "*Top 10 Scorer of The month:*\n\n"
    for idx, (user_id, username, score,round) in enumerate(top_10, 1):
        message += f"{idx}: @{escape_markdown(str(username))} \nScore: {escape_markdown(str(score))}      Rounds: {round} \n\n"
    try:
        await update.message.reply_text(message, parse_mode='MarkdownV2')
    except telegram.error.BadRequest:
        await update.message.chat.send_message(message, parse_mode='MarkdownV2')

# Command to show the user's rank and score
async def my_rank(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    username = update.message.from_user.username or update.message.from_user.first_name

    scores = load_scores()

    if not scores:
        try:
            await update.message.reply_text("No scores found")
        except telegram.error.BadRequest:
            await update.message.chat.send_message("No scores found")
        return

    # Sort by score in descending order
    scores.sort(key=lambda x: x[2], reverse=True)

    # Find user's rank
    user_rank = None

    for rank, (u_id, u_name, score,round) in enumerate(scores, 1):
        
        if str(u_id) == str(user_id):
            user_rank = (rank, score, round)
            break

    if user_rank:
        rank, score,round = user_rank
        try:
            await update.message.reply_text(f"Your rank: {rank}\nYour score: {score} in {round} round")
        except telegram.error.BadRequest:
            await update.message.chat.send_message(f"Your rank: {rank}\nYour score: {score} in {round} round")
    else:
        try:
            await update.message.reply_text("You haven't played the game yet")
        except telegram.error.BadRequest:
            await update.message.chat.send_message("You haven't played the game yet")

# Function to get a random word from the Excel file

async def start_game_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        global is_quiz_active, correct_users, chat_id, unanswered_poll,cancel_active,selected_quizscore_count,quiz_kick
        cancel_active = False
        quiz_kick= False
        
        reset_used_srnos()
        chat_id = update.message.chat.id
        
        if chat_id not in ALLOWED_GROUP_IDS:
            try:
                
                
                await update.message.reply_text("To Make your Own Bot and Start The Quiz In Your Group Talk to the Bot Creater @O000000000O00000000O")
            except (BadRequest, Forbidden, TimedOut) as e:
                await update.message.chat.send_message("To Make your Own Bot and Start The Quiz In Your Group Talk to the Bot Creater @O000000000O00000000O")
            return
        # Check if a quiz is already active
        if is_quiz_active:
            try:
                await update.message.chat.send_message("A quiz is already running. Please wait for it to finish before starting a new one. or use /cancelquiz")
            except (BadRequest, Forbidden, TimedOut) as e:
                print(f"Error canceling the quiz: {e}")
                
            return

        is_quiz_active = True  # Set to True when a new quiz starts
        quiz_state = {}
        selected_quizscore_count=0
        
        correct_users.clear()  # Reset scores at the beginning of each new quiz

        difficulty_keyboard = [
            [InlineKeyboardButton("NDA-CDS", callback_data='type_NDA0')],
            [InlineKeyboardButton("English Grammar", callback_data='type_BASIC')],
        ]
        reply_markup = InlineKeyboardMarkup(difficulty_keyboard)
        try:
            await update.message.chat.send_message('Select the Quiz type:', reply_markup=reply_markup)
        except (BadRequest, Forbidden, TimedOut) as e:
                print(f"Error canceling the quiz: {e}")
    except (BadRequest, Forbidden, TimedOut) as e:
                print(e)
#================================================================Option buttons For Quiz Type 
def Nda_keyboard0():
    return [
        [InlineKeyboardButton("Synonyms", callback_data='difficulty_synonyms_nda')],
        [InlineKeyboardButton("Antonyms", callback_data='difficulty_nda_antonyms')],
        [InlineKeyboardButton("Idiom-Phrase", callback_data='difficulty_idiomphrase_nda')],
        [InlineKeyboardButton("One word Substitute", callback_data='difficulty_nda_ows')],
        [InlineKeyboardButton("🧑‍🦯‍➡️ Next 🧑‍🦯‍➡️", callback_data='type_NDA1')]
    ]        

def Nda_keyboard1():
    return [
        
        [InlineKeyboardButton("Active-passive", callback_data='difficulty_acitvepassive_nda')],
        [InlineKeyboardButton("Fill in the Blanks", callback_data='difficulty_fillblank_nda')],
        [InlineKeyboardButton("Sentence Arrangement", callback_data='difficulty_nda_sentenceArrange')], 
        [InlineKeyboardButton("🏎️  Back ", callback_data='type_NDA0'),InlineKeyboardButton("Next 🧑‍🦯‍➡️", callback_data='type_NDA0')],
        
    ]
def Nda_keyboard2():
    return [
        [InlineKeyboardButton("Reasoning", callback_data='difficulty_nda_reasoning')],
        [InlineKeyboardButton("Physics-Chem-bio", callback_data='difficulty_nda_pcb')],
        [InlineKeyboardButton("Maths", callback_data='difficulty_nda_maths')],
        [InlineKeyboardButton("🏎️  Back", callback_data='type_NDA1'),InlineKeyboardButton("Next 🧑‍🦯‍➡️", callback_data='type_NDA0')]
        
    ]





#================================================================Option buttons For Quiz Type 
async def handle_type_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        query = update.callback_query
        username = query.from_user.username or query.from_user.first_name

        # If quiz is not active, ignore this input
        if not is_quiz_active:
            await query.answer("Please start a new quiz with /startquiz or cancel with /cancelquiz")
            return

        await query.answer()
        await query.answer()
        difficulty_message = ''

        if query.data == 'type_NDA0':
            selected_button_text = f"@{username} selected NDA-CDS Phase 1 \n Please wait..."
            try:
                await query.edit_message_text(text=selected_button_text)
            except (BadRequest, Forbidden, TimedOut) as e:
                print(f"Error canceling the quiz: {e}")
            
            reply_markup = InlineKeyboardMarkup(Nda_keyboard0())
            try:
                await query.message.chat.send_message('NDA Selected \n Select the Grammar Quiz type:', reply_markup=reply_markup)
            except (BadRequest, Forbidden, TimedOut) as e:
                print(f"Error canceling the quiz: {e}")
        elif query.data == 'type_NDA1':
            selected_button_text = f"@{username} selected NDA-CDS Phase 2 \n Please wait..."
            try:
                await query.edit_message_text(text=selected_button_text)
            except (BadRequest, Forbidden, TimedOut) as e:
                print(f"Error canceling the quiz: {e}")
            
            reply_markup = InlineKeyboardMarkup(Nda_keyboard1())
            try:
                await query.message.chat.send_message('NDA Selected \n Select the Grammar Quiz type:', reply_markup=reply_markup)
            except (BadRequest, Forbidden, TimedOut) as e:
                print(f"Error canceling the quiz: {e}")
        elif query.data == 'type_NDA2':
            selected_button_text = f"@{username} selected NDA-CDS Phase 3 \n Please wait..."
            try:
                await query.edit_message_text(text=selected_button_text)
            except (BadRequest, Forbidden, TimedOut) as e:
                print(f"Error canceling the quiz: {e}")
            
            reply_markup = InlineKeyboardMarkup(Nda_keyboard2())
            try:
                await query.message.chat.send_message('NDA Selected \n Select the Grammar Quiz type:', reply_markup=reply_markup)
            except (BadRequest, Forbidden, TimedOut) as e:
                print(f"Error canceling the quiz: {e}")
        elif query.data == 'type_BASIC':
            selected_button_text = f"@{username} Selected Basic English Grammar\n Please wait..."
            try:
                await query.edit_message_text(text=selected_button_text)
            except (BadRequest, Forbidden, TimedOut) as e:
                print(f"Error canceling the quiz: {e}")
            difficulty_keyboard = [
                [InlineKeyboardButton("Synonyms", callback_data='difficulty_synonyms')],
                [InlineKeyboardButton("Antonyms", callback_data='difficulty_antonyms')],
                [InlineKeyboardButton("Spelling Correction", callback_data='difficulty_spellcorr')],
                [InlineKeyboardButton("Sentence Correction", callback_data='difficulty_sentcorr')],
            ]
            reply_markup = InlineKeyboardMarkup(difficulty_keyboard)
            
            try:
                await query.message.chat.send_message('Basic Grammar Selected \n Select the Grammar Quiz type:', reply_markup=reply_markup)
            except (BadRequest, Forbidden, TimedOut) as e:
                print(f"Error canceling the quiz: {e}")
        elif query.data == 'type_AFCAT':
            selected_button_text = f"@{username} Selected AFCAT English Grammar\n Please wait..."
            try:
                await query.edit_message_text(text=selected_button_text)
            except (BadRequest, Forbidden, TimedOut) as e:
                print(f"Error canceling the quiz: {e}")
            difficulty_keyboard = [
                #[InlineKeyboardButton("Synonyms", callback_data='difficulty_synonyms')],
                #[InlineKeyboardButton("Antonyms", callback_data='difficulty_antonyms')],
                [InlineKeyboardButton("Spelling Correction", callback_data='difficulty_spellcorr')],
                #[InlineKeyboardButton("Sentence Correction", callback_data='difficulty_sentcorr')],
            ]
            reply_markup = InlineKeyboardMarkup(difficulty_keyboard)
            
            try:
                await query.message.chat.send_message('AFCAT Grammar Selected \n Select the Grammar Quiz type:', reply_markup=reply_markup)
            except (BadRequest, Forbidden, TimedOut) as e:
                print(f"Error canceling the quiz: {e}")
        elif query.data == 'type_CGL':
            selected_button_text = f"@{username} Selected SSC CGL English Grammar\n Please wait..."
            try:
                await query.edit_message_text(text=selected_button_text)
            except (BadRequest, Forbidden, TimedOut) as e:
                print(f"Error canceling the quiz: {e}")
            difficulty_keyboard = [
                #[InlineKeyboardButton("Synonyms", callback_data='difficulty_synonyms')],
                [InlineKeyboardButton("Antonyms", callback_data='difficulty_antonyms')],
                #[InlineKeyboardButton("Spelling Correction", callback_data='difficulty_spellcorr')],
                #[InlineKeyboardButton("Sentence Correction", callback_data='difficulty_sentcorr')],
            ]
            reply_markup = InlineKeyboardMarkup(difficulty_keyboard)
            
            try:
                await query.message.chat.send_message('SSC CGL Grammar Selected \n Select the Grammar Quiz type:', reply_markup=reply_markup)
            except (BadRequest, Forbidden, TimedOut) as e:
                print(f"Error canceling the quiz: {e}")
        
    except (BadRequest, Forbidden, TimedOut) as e:
                print(e)
          
# Handle difficulty selection
async def handle_difficulty_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        global EXCEL_FILE, Quiz_grammar_type
        query = update.callback_query
        username = query.from_user.username or query.from_user.first_name

        # If quiz is not active, ignore this input
        if not is_quiz_active:
            await query.answer("Please start a new quiz with /startquiz or cancel with /cancelquiz")
            return

        await query.answer()
        difficulty_message = ''

        if query.data == 'difficulty_synonyms':
            EXCEL_FILE = 'SYNO5.xlsx'
            difficulty_message = "Synonyms"
        elif query.data == 'difficulty_antonyms':
            EXCEL_FILE = 'Antonyms5.xlsx'
            difficulty_message = "Antonyms"
        elif query.data == 'difficulty_spellcorr':
            EXCEL_FILE = 'spellCorrection4.xlsx'
            difficulty_message = "Spelling Correction"
        elif query.data == 'difficulty_sentcorr':
            EXCEL_FILE = 'sentenceCorr4.xlsx'
            difficulty_message = "Sentence Correction"
        #==========================For NDA==============================
        elif query.data == 'difficulty_synonyms_nda':
            EXCEL_FILE = 'NDA_Synonyms1.xlsx'
            difficulty_message = "Synonyms"
        elif query.data == 'difficulty_acitvepassive_nda':
            EXCEL_FILE = 'NDA_active_passive_voice1.xlsx'
            difficulty_message = "Active passive Voice"
        elif query.data == 'difficulty_fillblank_nda':
            EXCEL_FILE = 'NDA_fillblank1.xlsx'
            difficulty_message = "Fill in the blanks"
        elif query.data == 'difficulty_idiomphrase_nda':
            EXCEL_FILE = 'NDA_idiom_phrase1.xlsx'
            difficulty_message = "Idiom Phrase"
        elif query.data == 'difficulty_nda_sentenceArrange':
            EXCEL_FILE = 'Nda_SentenceArrangement1_updated.xlsx'
            difficulty_message = "Sentence Arrangement"
        elif query.data == 'difficulty_nda_ows':
            EXCEL_FILE = 'Nda_1onewordsubstitute_updated.xlsx'
            difficulty_message = "One word Substitution"
        elif query.data == 'difficulty_nda_antonyms':
            EXCEL_FILE = 'Nda_Antonyms_updated.xlsx'
            difficulty_message = "Antonyms"
        elif query.data == 'difficulty_nda_reasoning':
            EXCEL_FILE = 'Nda_Reasoning1_updated.xlsx'
            difficulty_message = "Reasoning"
        elif query.data == 'difficulty_nda_pcb':
            EXCEL_FILE = 'Nda_PCB_hindi1_updated.xlsx.xlsx'
            difficulty_message = "Physics-Chem-bio"
        elif query.data == 'difficulty_nda_maths':
            EXCEL_FILE = 'Nda_Maths_updated.xlsx'
            difficulty_message = "Maths"

        # Edit the message to indicate selection and remove other buttons
        Quiz_grammar_type = difficulty_message
        selected_button_text = f"@{username} Chooses the {difficulty_message} For this round. Please wait... \n It Is Mandatory To Vote On Last Quiz"
        try:
            await query.edit_message_text(text=selected_button_text)
        except (BadRequest, Forbidden, TimedOut) as e:
            print(f"Error canceling the quiz: {e}")
            
            

        # Proceed with time limit selection
        if(Quiz_grammar_type !='Reasoning' and Quiz_grammar_type !='Maths'):
            time_keyboard = [
            [InlineKeyboardButton("10 Seconds", callback_data='time_10')],
            [InlineKeyboardButton("15 Seconds", callback_data='time_15')],
            [InlineKeyboardButton("20 Seconds", callback_data='time_20')],
            [InlineKeyboardButton("30 Seconds", callback_data='time_30')],
            
            ]
        else:
            time_keyboard = [
            [InlineKeyboardButton("30 Seconds", callback_data='time_30')],
            [InlineKeyboardButton("45 Seconds", callback_data='time_45')],
            [InlineKeyboardButton("60 Seconds", callback_data='time_60')],
            [InlineKeyboardButton("90 Seconds", callback_data='time_90')],
            ]
        reply_markup = InlineKeyboardMarkup(time_keyboard)
        try:
            await query.message.chat.send_message(f"{difficulty_message}. Select the time limit for each poll:", reply_markup=reply_markup)
        except (BadRequest, Forbidden, TimedOut) as e:
            print(f"Error canceling the quiz: {e}")
    except (BadRequest, Forbidden, TimedOut) as e:
                print(e)
        


# Handle time selection
async def handle_time_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        global selected_time_limit
        query = update.callback_query
        username = query.from_user.username or query.from_user.first_name

        # If quiz is not active, ignore this input
        if not is_quiz_active:
            await query.answer("Please start a new quiz with /startquiz or cancel with /cancelquiz")
            return
          
        await query.answer()

        # Map callback data to actual time values
        time_mapping = {
            'time_10': 4,
            'time_15': 15,
            'time_20': 20,
            'time_30': 30,
            'time_45': 45,
            'time_60': 60,
            'time_90': 90,
        }
        selected_time_limit = time_mapping.get(query.data, 10)
        selected_time_text = f"@{username} selected {selected_time_limit} second To complete one quiz. "

        # Edit the message to indicate time selection and remove other buttons
        try:
            await query.edit_message_text(text=selected_time_text)
        except (BadRequest, Forbidden, TimedOut) as e:
                print(f"Error canceling the quiz: {e}")
                

        # Round selection buttons
        keyboard = [
            [InlineKeyboardButton("15 Words", callback_data='15')],
            [InlineKeyboardButton("25 Words", callback_data='25')],
            [InlineKeyboardButton("35 Words", callback_data='35')],
            [InlineKeyboardButton("50 Words", callback_data='50')],
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        try:
            await query.message.chat.send_message(f"{selected_time_text}. How many rounds?", reply_markup=reply_markup)
        except (BadRequest, Forbidden, TimedOut) as e:
                print(f"Error canceling the quiz: {e}")
    except (BadRequest, Forbidden, TimedOut) as e:
                print(e)
            
async def cancel_quiz_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        if cancel_active:
            return None

        global is_quiz_active, quiz_state, correct_users
        msg_id = update.message.chat.id
        if msg_id not in ALLOWED_GROUP_IDS:
            try:
                await update.message.reply_text("To Make your Own Bot and Start The Quiz In Your Group Talk to the Bot Creater @O000000000O00000000O    ")
            except (BadRequest, Forbidden, TimedOut) as e:
                await update.message.chat.send_message("To Make your Own Bot and Start The Quiz In Your Group Talk to the Bot Creater @O000000000O00000000O    ")
            return
        # Check if the quiz is active
        chat_id = update.message.chat.id
        if not is_quiz_active:
            try:
                await update.message.chat.send_message("No quiz is currently active. \n To start This Quiz Write or Click /startquiz")
            except (BadRequest, Forbidden, TimedOut) as e:
                print(f"Error canceling the quiz: {e}")
            return
        
        # Reset global variables related to the quiz
        is_quiz_active = False
        quiz_state.clear()
        

        # Notify users that the quiz has been canceled
        try:
            await update.message.chat.send_message("The quiz has been canceled and reset. \n To start This Quiz Again Write or Click /startquiz")
        except (BadRequest, Forbidden, TimedOut) as e:
                print(f"Error canceling the quiz: {e}")
    except (BadRequest, Forbidden, TimedOut) as e:
                print(e)
# Display quiz results

# Handle button click and start quizzes
async def handle_button_click(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        global selected_poll_count,active_poll,cancel_active,quiz_kick
        query = update.callback_query
        username = query.from_user.username or query.from_user.first_name
      
        # If quiz is not active, ignore this input
        if not is_quiz_active:
            await query.answer("Please start a new quiz with /startquiz or cancel with /cancelquiz")
            return

        selected_poll_count = 5 #int(query.data)
        active_poll = selected_poll_count
        selected_rounds_text = f"@{username} selected {selected_poll_count} rounds. Starting the quiz..."

        # Edit the message to indicate round selection and remove other buttons
        try:
            await query.edit_message_text(text=selected_rounds_text)
        except (BadRequest, Forbidden, TimedOut) as e:
                print(f"Error canceling the quiz: {e}")
            

        cancel_active = True
        selected_polls = load_quiz_data(EXCEL_FILE,selected_poll_count) 
        selected_polls.append({
            'question': "Do You Want The Result of The Quiz",
            'options': ["yes", "of course", "why not", "yupp"],
            'correct_answer': "yes",
            'meaning': "To Show the result it is mandatory to Click on the Last poll \n Result Will Be Shown within 15 seconds"
        })
        
        
        for i, poll in enumerate(selected_polls):

            if quiz_kick:
                break
            try:
                poll_message = await context.bot.send_poll(
                chat_id=chat_id, 
                question=f"{i+1}/{selected_poll_count}: {poll['question']}",
                options=poll['options'],
                is_anonymous=False,
                allows_multiple_answers=False,
                type=Poll.QUIZ,
                correct_option_id=poll['options'].index(poll['correct_answer'])
                )
            except (BadRequest, Forbidden, TimedOut) as e:
                print(e)

            # Store the poll details in quiz_state
            quiz_state[poll_message.poll.id] = {
                "chat_id": chat_id,
                "question": poll["question"],
                "correct_answer": poll["correct_answer"],
                "options": poll["options"],
                "meaning": poll["meaning"], 
                "responses": {},
                "poll_number": i + 1,
                "expiry_time": time.time() + selected_time_limit,
                "poll_message": poll_message,
                "response_count": 0, 
                "users": [], 
            }

            # Start countdown and close poll
            await countdown_and_close_poll(poll_message, selected_time_limit, context)
            await asyncio.sleep(1)
    except (BadRequest, Forbidden, TimedOut) as e:
        print(e) 

# Countdown and close poll after time expires, with sending meaning
async def countdown_and_close_poll(poll_message, countdown_time, context):
    try:
        global quiz_kick
        # Wait for the countdown time to pass
        await asyncio.sleep(countdown_time)
        print("enter here 1")
        try:
            # Stop the poll after the time limit expires
            await poll_message.stop_poll()
        except Forbidden:
            # If the bot was kicked from the group, reset the quiz state for the chat
            chat_id = poll_message.chat.id
            global is_quiz_active
            quiz_kick = True

            is_quiz_active = False  # Mark the quiz as inactive
            quiz_state.clear()  # Clear the quiz state
            correct_users.clear()  # Reset user scores

            # Notify users that the bot was kicked and quiz is reset
            try:
                await context.bot.send_message(
                    chat_id=chat_id, 
                    text="The bot was kicked from the group. The quiz has been canceled and reset. You can start a new quiz after re-adding the bot."
                )
            except Exception as e:
                print(f"Error sending message: {e}")

            
            return

        # If the poll stops successfully, proceed to get the meaning and other actions
        poll_id = poll_message.poll.id
        if poll_id not in quiz_state:
            return

        # Get the quiz data for this poll
        quiz_data = quiz_state[poll_id]
        #meaning = quiz_data["meaning"]  # Retrieve the meaning from quiz data
        meaning = " ".join(str(quiz_data["meaning"]).split())
        print("this is a meaning",meaning)                  
        if((meaning != 'nan') and (meaning!='') and (meaning != 'No meaning provided')):             
            try:                 
                await context.bot.send_message(chat_id=quiz_data["chat_id"], text=f"Meaning: {meaning}")             
            except (BadRequest, Forbidden, TimedOut) as e:                 
                print(f"Error canceling the quiz: {e}")

        # Add a small delay before proceeding to the next poll
        await asyncio.sleep(1)
    except (BadRequest, Forbidden, TimedOut) as e:
        print(e)

final_poll_responses = {}

# Handle poll answers
async def handle_poll_answer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        global final_poll_responses, display_chat, selected_quizscore_count, correct_users

        answer: PollAnswer = update.poll_answer
        poll_id = answer.poll_id
        user_id = str(answer.user.id)  # Ensure user_id is always treated as a string
        username = answer.user.username or answer.user.first_name or user_id
        selected_options = answer.option_ids

        # Check if poll ID exists in quiz_state
        if poll_id not in quiz_state:
            return

        # Get the quiz data for the poll
        quiz_data = quiz_state[poll_id]
        correct_answer = quiz_data["correct_answer"]
        options = quiz_data["options"]
        curr_poll = quiz_data["poll_number"]
        display_chat = quiz_data["chat_id"]
        quiz_data["response_count"] += 1

        # Get the user's selected answer
        selected_answer = options[selected_options[0]]  # Assuming single choice

        # Store the user's response temporarily (no scoring yet)
        quiz_data["responses"][user_id] = selected_answer

        # Track correct answers
        if selected_answer == correct_answer:
            if curr_poll == selected_poll_count + 1:
                print("skip this part")
            else:
                # Initialize or update the user's score in correct_users
                if user_id not in correct_users:
                    correct_users[user_id] = {"username": username, "score": 0}
                correct_users[user_id]["score"] += 1  # Increment the score
                
        if user_id not in quiz_data["users"]:
            quiz_data["users"].append(user_id)

        # Handle the last poll
        if curr_poll == selected_poll_count + 1:
            print("equal equal", selected_quizscore_count)
            await asyncio.sleep(5)
            if selected_quizscore_count == 0:
                print("equal equal 2", selected_quizscore_count)
                await calculate_scores(update, context)
                update_user_score(correct_users)
                final_poll_responses = {}
                selected_quizscore_count = 1

    except (BadRequest, Forbidden, TimedOut) as e:
        print(e)

async def download_scores_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        chat_id = update.message.chat.id
        
        # Check if the chat_id (group ID) is in the allowed list
    
        if chat_id not in ALLOWED_GROUP_IDS:
            try:
                await update.message.reply_text("Due to the free service, you are not allowed to start a game in this group. Play there https://t.me/+yVFKtplWZUA0Yzhl or contact @O000000000O00000000O")
            except telegram.error.BadRequest:
                await update.message.chat.send_message("Due to the free service, you are not allowed to start a game in this group. Play there https://t.me/+yVFKtplWZUA0Yzhl or contact @O000000000O00000000O")
            return

        # Check if the file exists
        if os.path.exists(SCORE_FILE):
            # Send the file to the user
            with open(SCORE_FILE, 'rb') as file:
                await context.bot.send_document(chat_id=update.message.chat.id, document=file)
        else:
            # Notify the user that the file does not exist
            await update.message.reply_text("Sorry, the score file is not available.")
    except Exception as e:
        # Handle any errors
        await update.message.reply_text(f"An error occurred: {e}")

def update_user_score(correct_users):
    """
    Update user scores in an Excel file. Add new users if they don't exist.
    """
    try:
        game_round = 1
        # Load existing workbook or create a new one
        try:
            workbook = load_workbook(SCORE_FILE)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Scores"
            sheet.append(["Sr No", "User ID", "Username", "Score", "Round"])

        # Load existing scores into a dictionary
        existing_scores = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row:  # Ensure row is not empty
                if len(row) == 5:  # If all 5 columns are present
                    sr_no, user_id, username, score, round = row
                elif len(row) == 4:  # Missing "Round" column, initialize round to 0
                    sr_no, user_id, username, score = row
                    round = 0
                else:
                    continue  # Skip invalid rows

                existing_scores[str(user_id)] = {
                    "sr_no": sr_no,
                    "username": username,
                    "score": int(score),
                    "round": int(round)
                }

        # Update scores based on correct_users
        for user_id, data in correct_users.items():
            username = data["username"]
            new_score = data["score"]

            if user_id in existing_scores:
                # Update existing user's score
                existing_scores[user_id]["score"] += new_score
                existing_scores[user_id]["round"] += game_round
            else:
                # Add new user
                sr_no = len(existing_scores) + 1
                existing_scores[user_id] = {
                    "sr_no": sr_no,
                    "username": username,
                    "score": new_score,
                    "round": game_round
                }

        # Clear existing rows and rewrite updated scores
        sheet.delete_rows(2, sheet.max_row)
        for user_id, data in existing_scores.items():
            sheet.append([data["sr_no"], user_id, data["username"], data["score"], data["round"]])

        # Save workbook
        workbook.save(SCORE_FILE)
        print("Scores updated successfully.")

    except Exception as e:
        print(f"Error updating scores: {e}")

# Function to calculate scores
async def calculate_scores(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        # Check if there are any responses in `correct_users` dictionary to display results
        await display_results(update, context)
        
    except (BadRequest, Forbidden, TimedOut) as e:
        print(e)

# Display quiz results, even if only partial or no responses are available
async def display_results(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global is_quiz_active, Quiz_grammar_type, correct_users
    cancel_active = False

    if display_chat:
        chatid = display_chat
    else:
        chatid = chat_id  # Ensure `chatid` is used correctly here.

    # Check if there are any scores to display
    try:
        result_message = (
            f"🎉 Quiz Results: 🥳🥳🥳🥳\nIn {Quiz_grammar_type} Out Of {selected_poll_count}\n"
            "Here are the top scorers of This Round:\n\n"
        )

        # Sort by the number of correct answers and display each username with their score
        sorted_results = sorted(
            correct_users.items(),
            key=lambda x: x[1]["score"],  # Sort by the "score" field
            reverse=True
        )

        top_10_results = sorted_results[:10]

        if not top_10_results:
            result_message = "No scores received. No one answered correctly."
        else:
            p = 1
            for user_id, user_data in top_10_results:
                username = user_data["username"]
                score = user_data["score"]

                if p == 1:
                    result_message += f"🏆)- @{username}: {score}\n"
                elif p == 2:
                    result_message += f"🥈)- @{username}: {score}\n"
                elif p == 3:
                    result_message += f"🥉)- @{username}: {score}\n"
                else:
                    result_message += f"🧌{p})- @{username}: {score}\n"
                p += 1

        result_message += "\nTo start this quiz again, write or click /startquiz"

        # Send the results message to the chat
        try:
            await context.bot.send_message(chat_id=chatid, text=result_message)
        except Exception as e:
            print(f"Error sending message: {e}")

    except Exception as e:
        print(f"Error preparing results message: {e}")

    # Reset quiz active state
    is_quiz_active = False

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    help_message = (
        "If you need any kind of help or have suggestions, please discuss with my owner From Description"
    )
    await update.message.chat.send_message(help_message)


def main():
    application = Application.builder().token(TOKEN).build()
    application.add_handler(CommandHandler('startquiz', start_game_command))
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(CallbackQueryHandler(handle_difficulty_selection, pattern='^difficulty_'))
    application.add_handler(CallbackQueryHandler(handle_type_selection, pattern='^type_'))
    application.add_handler(CallbackQueryHandler(handle_time_selection, pattern='^time_'))
    application.add_handler(CallbackQueryHandler(handle_button_click, pattern=r'^\d+$'))
    application.add_handler(PollAnswerHandler(handle_poll_answer))
    application.add_handler(CommandHandler('cancelquiz', cancel_quiz_command))
    application.add_handler(CommandHandler('myrank', my_rank))
    application.add_handler(CommandHandler('top10score', select_top_10_users))
    application.add_handler(CommandHandler('deleteuserscores0404', delete_user_scores))
    application.add_handler(CommandHandler('downloadscoreiesp', download_scores_command))

    # Start the bot
    application.run_polling()

if __name__ == '__main__':
    main()
